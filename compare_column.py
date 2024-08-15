# encoding: utf-8
# @Time    : 2024/08/13
# @Author  : 半只程序员
# @Email   : 18152007693@163.com
# @File    : compare_column.py
# @Software: PyCharm

from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from .logger import logger
from .read_table import get_data_from_file


def compare_two_column(data, column1, column2, output):
    """
    比较两个列，如果column1的值在column2中存在，则将column1的值标为红色
    :param data: 源数据
    :param column1:字段1
    :param column2:字段2
    :param output:对比后的输出文件
    :return:msg
    """
    try:
        df = get_data_from_file(data)
    except UnicodeDecodeError as e:
        df = get_data_from_file(data, encoding='utf-16')
    try:
        wb = Workbook()
        ws = wb.active
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx > 1:  # Skip header
                    if df.columns[c_idx - 1] == column1 and value in df[column2].values:
                        cell.fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        wb.save(output)
        msg = '字段`{}`和`{}`对比完成，请在{}中查看对比结果，红色背景为相同数据'.format(column1, column2, output)
        # msg输出到控制台
        logger.info(msg)
    except PermissionError as e:
        msg = '文件{}被占用，请关闭后再运行'.format(output)
        logger.error(msg)


if __name__ == '__main__':
    compare_two_column(r'C:\Users\YAFEX\Desktop\inflationStatement\version\出入库明细改造\files\FFGYS_2024-01.csv',
                       'inventoryTaxedPriceRmb', 'inventoryUntaxedAmount', 'FFGYS_2024-01.xlsx')
